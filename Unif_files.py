import time

import openpyxl
import os
import win32com.client
import glob
import string
import openpyxl
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import _get_column_letter
from openpyxl.worksheet.cell_range import CellRange

letter = string.ascii_letters
PATH_File_info = "C:/Users/Kobi Malul/Desktop/Fix File/Fix File.xlsx"
folder_unif = "C:/Users/Kobi Malul/Desktop/LOG/FilesToLoad"
end_file_want = [".xlsx",".XLSX"]
o = win32com.client.Dispatch("Excel.Application")
o.Visible = False
input_dir = r"C:\Users\Kobi Malul\Desktop\LOG\FilesToLoad"
output_dir = r"C:\Users\Kobi Malul\Desktop\LOG\FilesToLoad"
files_xls = glob.glob(input_dir + "/*.xls")
date = "20211231"
unif_folder = "C:/Users/Kobi Malul/Desktop/LOG/Unif"




class Delete_empty_row:
    def __init__(self):
        self.open_info = PATH_File_info
        self.files_on_folder = folder_unif
        self.open_file = openpyxl.load_workbook(self.open_info)
        self.active_open_file = self.open_file.active
        self.get_max_row = self.active_open_file.max_row
        self.get_max_col = self.active_open_file.max_column

    def Change_Type(self):
        for files in os.listdir(folder_unif):
            num_slice_number_kopa = files.find("_")
            self.number_kopa = files[0:num_slice_number_kopa]
            print(self.number_kopa)
            filename_end = os.path.splitext(files)[-1]
            if filename_end == ".xls":
                for filename in files_xls:
                        try:
                            if "162_orig_20210930.xls" in filename and "1162_orig_20210930.xls" not in filename:
                                print("קופץ חלון לשינוי שם באקסל נא לרשום סתם משהו עליך למזער את כל הקבצים הפתוחים לפתוח את התוכנה ולמזער שוב")
                            print(filename)
                            file = os.path.basename(filename)
                            output = output_dir + '/' + file.replace('.xls', '.xlsx')
                            wb = o.Workbooks.Open(filename)
                            wb.ActiveSheet.SaveAs(output, 51)
                            wb.Close(True)
                            Delete_empty_row.delete_file(self)
                        except:
                            print(filename)
            elif filename_end in end_file_want:
                Delete_empty_row.delete_file(self)

    def delete_file(self):
        for file in os.listdir(folder_unif):
            end_file = os.path.splitext(file)[-1]
            if end_file == ".xls":
                change_end = file.replace(".xls",".xlsx")
                if change_end in os.listdir(folder_unif):
                    os.remove(folder_unif + "/" + file)
            elif end_file in end_file_want:
                self.file_active_folder = file
                self.open_file_to_delete_rows = openpyxl.load_workbook(folder_unif + "/" + self.file_active_folder)
                self.new_file_name = self.file_active_folder.replace("_orig","")
                self.open_file_to_delete_rows.save(unif_folder + "/" + self.new_file_name)
                Delete_empty_row.get_info(self)

    def get_info(self):
        for col in range(0,self.get_max_col):
            for rows in range(1,self.get_max_row):
                self.select_first_sheet = self.open_file["שיוך קופה"]
                self.search_kopa = str(self.select_first_sheet["{}{}".format(letter[col],rows)].value)
                if self.number_kopa == self.search_kopa:
                    self.mtafel = str(self.select_first_sheet["{}{}".format(letter[col+1],rows)].value)
                    self.select_sheet = str(self.select_first_sheet["{}{}".format(letter[col+2],rows)].value)
                    Delete_empty_row.get_sheet(self)

    def get_sheet(self):
        self.sheets = []
        self.select_unif_sheet = self.open_file[self.select_sheet]
        for col_unif in range(0, self.get_max_col):
            for row_unif in range(2, self.get_max_row):
                self.search_mtafel = self.select_unif_sheet["{}{}".format(letter[col_unif], row_unif)].value
                self.sheet = self.select_unif_sheet["B{}".format(row_unif)].value
                if self.search_mtafel == self.mtafel and self.sheet not in self.sheets:
                    self.sheet = self.select_unif_sheet["B{}".format(row_unif)].value
                    print(self.sheet)
                    self.sheets.append(self.sheet)
                    Delete_empty_row.unif(self)

    def unif(self):
            self.first_row_delete = 0
            self.flag_start_row = 0
            self.flag = 0
            for col_unif in range(0,self.get_max_col):
                for row_unif in range(2, self.get_max_row):
                    self.search_mtafel = self.select_unif_sheet["{}{}".format(letter[col_unif],row_unif)].value
                    self.value_of_row = self.select_unif_sheet["E{}".format(row_unif)].value
                    self.sheet_right_now = self.select_unif_sheet["B{}".format(row_unif)].value
                    self.Value_place = self.select_unif_sheet["C{}".format(row_unif)].value
                    self.value_need_to_be = self.select_unif_sheet["D{}".format(row_unif)].value
                    if self.search_mtafel == self.mtafel and self.Value_place!= None and self.sheet_right_now == self.sheet:
                        Delete_empty_row.Value_be(self)

                    if self.search_mtafel == self.mtafel and self.value_of_row != None and self.flag != 1 and self.sheet_right_now == self.sheet:
                        self.first_row_delete = self.value_of_row
                        self.last_row_delete = self.select_unif_sheet["F{}".format(row_unif)].value
                        self.flag = 1
                    elif self.first_row_delete > 0 and self.value_of_row != None and self.value_of_row >self.first_row_delete:
                        self.seconed_time_delete = self.value_of_row
                        self.seconed_time_delete_list_row = self.select_unif_sheet["F{}".format(row_unif)].value

            if self.first_row_delete > 0 or self.value_of_row != None and self.seconed_time_delete> 0:
                Delete_empty_row.delete_rows_unif(self)

    def Value_be(self):
        self.open_for_delete_rows_sheets = openpyxl.load_workbook(unif_folder + "/" + self.new_file_name)
        self.active_to_delete = self.open_for_delete_rows_sheets.active
        self.select_sheet_unif_file = self.open_for_delete_rows_sheets[self.sheet]
        self.select_sheet_unif_file["{}".format(self.Value_place)].value = str(self.value_need_to_be)
        self.open_for_delete_rows_sheets.save(unif_folder + "/" + self.new_file_name)

    def delete_rows_unif(self):
        self.open_for_delete_rows_sheets = openpyxl.load_workbook(unif_folder + "/" + self.new_file_name)
        self.active_to_delete = self.open_for_delete_rows_sheets.active
        self.select_sheet_unif_file = self.open_for_delete_rows_sheets[self.sheet]
        self.sheetspaly = self.open_for_delete_rows_sheets.sheetnames
        for sheeets_unif in self.sheetspaly:
            select_new_sheet = self.open_for_delete_rows_sheets[sheeets_unif]
            for merge in list(select_new_sheet.merged_cells):
                self.select_sheet_unif_file.unmerge_cells(range_string=str(merge))

        for first_rows_time in range(self.first_row_delete, self.last_row_delete+1):
            self.select_sheet_unif_file.delete_rows(idx=self.first_row_delete, amount=self.first_row_delete)
        for secund_time_delete in range(self.seconed_time_delete, self.seconed_time_delete_list_row):
            self.select_sheet_unif_file.delete_rows(idx=self.seconed_time_delete, amount=self.seconed_time_delete_list_row-1)

        self.open_for_delete_rows_sheets.save(unif_folder + "/" + self.new_file_name)
        Delete_empty_row.back_to_xls(self)

    def back_to_xls(self):
        self.active_open_delete = self.open_for_delete_rows_sheets.active
        self.sheet_name_delete = self.open_for_delete_rows_sheets.sheetnames
        files_unif = r"C:\Users\Kobi Malul\Desktop\LOG\Unif"
        output_unif = r"C:\Users\Kobi Malul\Desktop\LOG\Unif"
        files_unif = glob.glob(files_unif + "/*.xlsx")
        if len(self.sheet_name_delete)-1 <= len(self.sheets):
            for file_unif_xlsx in files_unif:
                    end_file_unif = os.path.splitext(file_unif_xlsx)[-1]
                    if end_file_unif in end_file_want:
                        file = os.path.basename(file_unif_xlsx)
                        output = output_unif + '/' + file.replace('.xlsx','.xls')
                        wb = o.Workbooks.Open(file_unif_xlsx)
                        wb.ActiveSheet.SaveAs(output,50)
                        wb.Close(True)
                        time.sleep(3)
                        Delete_empty_row.delete_file_xlsx(self)

    def delete_file_xlsx(self):
            for file in os.listdir(unif_folder):
                end_file = os.path.splitext(file)[-1]
                if end_file == ".xlsx":
                    print(file)
                    change_end = file.replace(".xlsx", ".xls")
                    if change_end in os.listdir(unif_folder):
                        os.remove(unif_folder + "/" + file)

unif_files = Delete_empty_row()
unif_files.Change_Type()
